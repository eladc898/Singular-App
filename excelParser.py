import datetime
import sys
from openpyxl import load_workbook


def cut_dollar_sign(string):
    return float(string[1:].replace(',', ''))


def parse_excel(file_path):
    # open the excel file and read the first sheet
    workbook = load_workbook(file_path, data_only=True)
    first_sheet = workbook.sheetnames[0]
    worksheet = workbook[first_sheet]

    total_installs = 0
    total_cost = 0
    data_by_date = {}
    data_by_app_and_platform = {}

    for row in worksheet.iter_rows():
        # save Totals according location in the file
        if row[4].column == 'E' and row[4].row == 61:
            total_cost = row[4].value
        if row[5].column == 'F' and row[5].row == 61:
            total_installs = row[5].value

        # escape first 5 rows and don't save the row of Totals in the dictionaries
        if row[0] is not None and row[0].row < 5 or not isinstance(row[0].value, datetime.datetime):
            continue

        # case the key (app,platform) already exists
        if (row[1].value, row[2].value) in data_by_app_and_platform:
            data_by_app_and_platform[(row[1].value, row[2].value)]['cost'] = \
                data_by_app_and_platform[(row[1].value, row[2].value)]['cost'] + cut_dollar_sign(row[4].value)
            data_by_app_and_platform[(row[1].value, row[2].value)]['install'] = \
                data_by_app_and_platform[(row[1].value, row[2].value)]['install'] + row[5].value
        # case of new key (app,platform)
        else:
            data_by_app_and_platform[(row[1].value, row[2].value)] = {'cost': cut_dollar_sign(row[4].value), 'install': row[5].value}
        # case the key date already exists
        if row[0].value in data_by_date:
            data_by_date[row[0].value]['cost'] = data_by_date[row[0].value].get('cost') + cut_dollar_sign(row[4].value);
            data_by_date[row[0].value]['install'] = data_by_date[row[0].value].get('install') + row[5].value;
        # case of new date
        else:
            data_by_date[row[0].value] = {'cost': cut_dollar_sign(row[4].value), 'install': row[5].value}

    print('The Total Cost:', total_cost)
    print('The Total Installs:', total_installs)
    # data by date:
    print('\ndata by date:\n')
    for k, v in data_by_date.items():
        print(k.strftime('%Y-%m-%d'), ':', 'cost:', v['cost'], 'install', v['install'])

    print('\ndata by app and platform:\n')
    for k, v in data_by_app_and_platform.items():
        print('App:', k[0], 'Platform:', k[1], ':', 'cost:', v['cost'], 'install', v['install'])


def find_file():
    if len(sys.argv) == 2 and sys.argv[1] is not None:
        file_path = sys.argv[1]
        print("This is the path of the excel: ", sys.argv[1])
        return file_path
    else:
        print("No file has passed as argument")
        exit(-1)


def main():
    file_path = find_file()
    parse_excel(file_path)


if __name__ == '__main__':
    main()
